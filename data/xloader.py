from openpyxl import load_workbook
from pymongo import MongoClient

import os
import sys
import urllib.parse

def extract_tasks(data_file):
    wb = load_workbook(filename=data_file, read_only=True)
    ws = wb.active
    tasks = []
    total_tasks = 0
    in_chapter = False
    last_lesson = None
    last_chapter = None
    for row in ws.rows:
        if "Grade" in repr(row[0].value):
            total_tasks += int(row[0].value.split(':')[2])
        if "Chapter" in repr(row[0].value):
            if "Test" not in row[0].value:
                last_chapter = row[0].value
                in_chapter = True
                continue
        if in_chapter:
            if row[0].value == "Worksheet":
                continue
            if "Lesson:" in repr(row[0].value):
                last_lesson = row[0].value
                continue
            elif "Week of:" in repr(row[0].value):
                in_chapter = False
            else:
                tasks.append({"task":row[0].value,
                              "lesson":last_lesson,
                              "chapter":last_chapter})
    wb.close()
    return tasks

def extract_tasks_by_week(data_file):
    wb = load_workbook(filename=data_file, read_only=True)
    ws = wb.active
    tasks = []
    working_rows = ws.rows
    working_rows_values = list(map(lambda x: x[0].value, working_rows))
    working_rows_values = list(filter(lambda y: y != None, working_rows_values))
    weeks = []
    week_boundries = find_value_bounds("Week", working_rows_values)
    for i in range(len(week_boundries) - 1):
        weeks.append(working_rows_values[week_boundries[i]+1:week_boundries[i+1]])
    for week in weeks:
        subject = week[0]
        grade = week[1].split("|")[0].strip()
        chapter = week[2]
        current_lesson = ""
        for item in week[3:]:
            if "Lesson:" in item:
                current_lesson = item
                continue
            else:
                tasks.append({
                "subject": subject,
                "lesson": current_lesson,
                "grade": grade,
                "chapter": chapter,
                "task": item
                })
    return tasks

def find_value_bounds(value, array):
    bounds = []
    for i in range(len(array)):
        if value in array[i]:
            bounds.append(i)
    return bounds

def split_by_bounds(bounds, array):
    split_arrays = []
    for i in range(len(bounds) - 1):
        split_arrays.append(array[bounds[i]:bounds[i+1]])
    split_arrays.append(array[bounds[-1]:])
    return split_arrays

def extract_with_preprocess(data_file):
    wb = load_workbook(filename=data_file, read_only=True)
    ws = wb.active
    tasks = []
    # get first collumn of each row
    working_rows = ws.rows
    working_rows_values = list(map(lambda x: x[0].value, working_rows))
    # remove all null values
    working_rows_values = list(filter(lambda y: y != None, working_rows_values))
    # extract repeated info then delete
    subject = working_rows_values[1]
    grade = working_rows_values[2].split("|")[0].strip()
    working_rows_values = list(filter(lambda y: y != subject and grade not in y and "Week of:" not in y, working_rows_values))
    # find all chapter bounds
    chapter_bounds = find_value_bounds('Chapter', working_rows_values)
    chapters = split_by_bounds(chapter_bounds, working_rows_values)
    # extract tasks with lesson info
    order = 0
    for chapter in chapters:
        current_lesson = ""
        for item in chapter[1:]:
            if "Lesson:" in item:
                current_lesson = item
                continue
            else:
                tasks.append({
                "subject": subject,
                "lesson": current_lesson,
                "grade": grade,
                "chapter": chapter[0],
                "task": item,
                "completed": False,
                "weight": 3 if ('lab' in item.lower()) else 1,
                "order": order
                })
                order += 1
    return tasks
